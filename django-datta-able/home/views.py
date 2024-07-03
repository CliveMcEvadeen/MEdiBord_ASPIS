from django.shortcuts import render, redirect
from admin_datta.forms import RegistrationForm, LoginForm, UserPasswordChangeForm, UserPasswordResetForm, UserSetPasswordForm
from django.contrib.auth.views import LoginView, PasswordChangeView, PasswordResetConfirmView, PasswordResetView
from django.views.generic import CreateView
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from .models import *
from turtle import clear
from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
import json
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import HttpResponse
from home import forms, models
from .models import *
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_protect
import shutil
import os
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect
import uuid
from django.http import JsonResponse
from django.urls import reverse
import pdb
import os
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import pandas as pd
from .forms import FileUploadForm, SchoolDetailsForm
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill,Border,Side


def index(request):

  context = {
    'segment'  : 'index',
    #'products' : Product.objects.all()
  }
  return render(request, "pages/index.html", context)

def tables(request):
  context = {
    'segment': 'tables'
  }
  return render(request, "pages/dynamic-tables.html", context)


# copie


context={
    'page':'',
    'page_title':'',
    'system_name':'Student Result Managament System',
    'short_name':'SRMS',
    'has_navigation':True,
    'has_sidebar':True,
}

# Create your views here.
@login_required
def home(request):
    context['page'] = 'home'
    context['page_title'] = 'Dashboard'
    context['classes'] = models.Class.objects.filter(status =1).count()
    context['subjects'] = models.Subject.objects.filter(status =1).count()
    context['students'] = models.Student.objects.filter(status =1).count()
    context['results'] = models.Result.objects.count()
    context['has_navigation'] = True
    context['has_sidebar'] = True
    return render(request,'home.html',context)

#login
def report():
    # renders the data to the front for population
    pass
def bulk_import(sheet_id):
    # allows the importation of data from the generated  sheet and consolidates the data
    # locks the sheet from being renamed
    # the sheet is protected
    pass

def generate_sheet(dir, sheet_name, sheet_id):
    # generates the datasheet that is stored in the system directory
    pass 

def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Incorrect username or password"
        else:
            resp['msg'] = "Incorrect username or password"
    return HttpResponse(json.dumps(resp),content_type='application/json')

#Logout
def logoutuser(student):
    logout(student)
    return redirect('/')


@login_required
def update_profile(request):
    context['page_title'] = 'Update Profile'
    user = User.objects.get(id = request.user.id)
    if not request.method == 'POST':
        form = forms.UpdateProfile(instance=user)
        context['form'] = form
        print(form)
    else:
        form = forms.UpdateProfile(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile has been updated")
            return redirect("profile-page")
        else:
            context['form'] = form
            
    return render(request, 'manage_profile.html',context)


@login_required
def update_password(request):
    context['page_title'] = "Update Password"
    if request.method == 'POST':
        form = forms.UpdatePasswords(user = request.user, data= request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Your Account Password has been updated successfully")
            update_session_auth_hash(request, form.user)
            return redirect("profile-page")
        else:
            context['form'] = form
    else:
        form = forms.UpdatePasswords(request.POST)
        context['form'] = form
    return render(request,'update_password.html',context)

@login_required
def profile(request):
    context['page'] = 'profile'
    context['page_title'] = "Profile"
    return render(request,'profile.html', context)


# Class
@login_required
def class_mgt(request):
    context['page'] = 'class_mgt'
    context['page_title'] = 'Class Management'
    classes = models.Class.objects.all()
    context['classes'] = classes
    return render(request,'class_mgt.html',context)


@login_required
def manage_class(request, pk = None):
    if not pk is None:
        classData = models.Class.objects.get(id = pk)
        context['classData'] = classData
    else:
        context['classData'] = {}
    return render(request, 'manage_class.html', context)

@login_required
def save_class(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        if POST['id'] == None or POST['id'] == '':
            form = forms.SaveClass(POST)
        else:
            _class = models.Class.objects.get(id = POST['id'])
            form = forms.SaveClass(POST, instance=_class)

        if form.is_valid():
            form.save()
            resp['status'] = 'success'
            messages.success(request, "Class Detail has been saved successfully.")
        else:
            resp['msg'] = 'Class Detail has failed to save.'
            for field in form:
                for error in field.errors:
                    resp['msg'] += str("<br/>"+error)
    return HttpResponse(json.dumps(resp),content_type="application/json")    

@login_required
def delete_class(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        try:
            models.Class.objects.get(id = POST['id']).delete()
            resp['status'] = 'success'
            messages.success(request, "Class Detail has been deleted successfully.")
        except:
            resp['msg'] = 'Class Detail has failed to delete.'

    return HttpResponse(json.dumps(resp),content_type="application/json")

#Subject
@login_required
def subject_mgt(request):
    context['page'] = 'subject_mgt'
    context['page_title'] = 'Subject Management'
    subjects = models.Subject.objects.all()
    context['subjects'] = subjects
    return render(request,'subject_mgt.html',context)


@login_required
def manage_subject(request, pk = None):
    if not pk is None:
        subject = models.Subject.objects.get(id = pk)
        context['subject'] = subject
    else:
        context['subject'] = {}
    return render(request, 'manage_subject.html', context)

@login_required
def view_subject(request, pk = None):
    if not pk is None:
        subject = models.Subject.objects.get(id = pk)
        context['subject'] = subject
    else:
        context['subject'] = {}
    return render(request, 'view_subject.html', context)

@login_required
def save_subject(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        if POST['id'] == None or POST['id'] == '':
            form = forms.SaveSubject(POST)
        else:
            subject = models.Subject.objects.get(id = POST['id'])
            form = forms.SaveSubject(POST, instance=subject)

        if form.is_valid():
            form.save()
            resp['status'] = 'success'
            messages.success(request, "Subject Detail has been saved successfully.")
        else:
            resp['msg'] = 'Subject Detail has failed to save.'
            for field in form:
                for error in field.errors:
                    resp['msg'] += str(f"<br/>"+error)
    return HttpResponse(json.dumps(resp),content_type="application/json")    

@login_required
def delete_subject(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        try:
            models.Subject.objects.get(id = POST['id']).delete()
            resp['status'] = 'success'
            messages.success(request, "Subject Detail has been deleted successfully.")
        except:
            resp['msg'] = 'Subject Detail has failed to delete.'

    return HttpResponse(json.dumps(resp),content_type="application/json")

#Students
@login_required
def student_mgt(request):
    context['page'] = 'student_mgt'
    context['page_title'] = 'Student Management'
    students = models.Student.objects.all()
    context['students'] = students
    return render(request,'student_mgt.html',context)


@login_required
def manage_student(request, pk = None):
    classes = models.Class.objects.filter(status = 1).all()
    context['classes'] = classes
    if not pk is None:
        student = models.Student.objects.get(id = pk)
        context['student'] = student
    else:
        context['student'] = {}
    return render(request, 'manage_student.html', context)

@login_required
def view_student(request, pk = None):
    if not pk is None:
        student = models.Student.objects.get(id = pk)
        context['student'] = student
    else:
        context['student'] = {}
    return render(request, 'view_student.html', context)

@login_required
def save_student(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        if POST['id'] == None or POST['id'] == '':
            form = forms.SaveStudent(POST)
        else:
            student = models.Student.objects.get(id = POST['id'])
            form = forms.SaveStudent(POST, instance=student)

        if form.is_valid():
            form.save()
            resp['status'] = 'success'
            messages.success(request, "Student Detail has been saved successfully.")
        else:
            resp['msg'] = 'Student Detail has failed to save.'
            for field in form:
                for error in field.errors:
                    resp['msg'] += str(f"<br/> [{field.name}] "+error)
    return HttpResponse(json.dumps(resp),content_type="application/json")    

@login_required
def delete_student(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        try:
            models.Student.objects.get(id = POST['id']).delete()
            resp['status'] = 'success'
            messages.success(request, "Student Detail has been deleted successfully.")
        except:
            resp['msg'] = 'Student Detail has failed to delete.'

    return HttpResponse(json.dumps(resp),content_type="application/json")

#Result
@login_required
def result_mgt(request):
    context['page'] = 'result_mgt'
    context['page_title'] = 'Result Management'
    results =models.Result.objects.all()
    context['results'] = results
    return render(request,'result_mgt.html',context)


@login_required
def manage_result(request, pk = None):
    students = models.Student.objects.filter(status = 1).all()
    context['students'] = students
    subjects = models.Subject.objects.filter(status = 1).all()
    context['subjects'] = subjects
    if not pk is None:
        result =models.Result.objects.get(id = pk)
        marks =models.Student_Subject_Result.objects.filter(result = result)
        context['marks'] = marks
        context['result'] = result
    else:
        context['result'] = {}
        context['marks'] = {}
    return render(request, 'manage_result.html', context)

def view_result(request, pk = None):
    if not pk is None:
        result =models.Result.objects.get(id = pk)
        context['result'] = result
        marks =models.Student_Subject_Result.objects.filter(result = result)
        context['marks'] = marks
    else:
        context['result'] = {}
        context['marks'] = {}
    return render(request, 'view_result.html', context)

@login_required
def save_result(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        if POST['id'] == None or POST['id'] == '':
            form = forms.SaveResult(POST)
        else:
            result =models.Result.objects.get(id = POST['id'])
            form = forms.SaveResult(POST, instance=result)
        if form.is_valid():
            form.save()
            is_new = False
            if POST['id'] == '':
                rid = models.Result.objects.all().last().id
                result =models.Result.objects.get(id = rid)
                is_new = True
            else:
                rid = POST['id']
            models.Student_Subject_Result.objects.filter(result = result).delete()
            has_error = False
            subjects= request.POST.getlist('subject[]')
            grade= request.POST.getlist('grade[]')
            i = 0
            for subject in subjects:
                data = {
                    'result' :rid,
                    'subject' :subject,
                    'grade' : grade[i]
                }
                form2 = forms.SaveSubjectResult(data = data)
                if form2.is_valid():
                    form2.save()
                else:
                    resp['msg'] = 'Result Detail has failed to save.'
                    for field in form2:
                        for error in field.errors:
                            resp['msg'] += str(f"<br/> [{field.name}] "+error)
                    has_error = True
                    break
                i +=1
            if has_error == False:
                resp['status'] = 'success'
                messages.success(request, "Result Detail has been saved successfully.")
            else:
                if is_new:
                    models.Result.objects.get(id = POST['id']).delete()
        else:
            resp['msg'] = 'Result Detail has failed to save.'
            for field in form:
                for error in field.errors:
                    resp['msg'] += str(f"<br/> [{field.name}] "+error)
    return HttpResponse(json.dumps(resp),content_type="application/json")    

@login_required
def delete_result(request):
    resp = { 'status':'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = 'Request has been sent without data.'
    else:
        POST = request.POST
        try:
            models.Result.objects.get(id = POST['id']).delete()
            resp['status'] = 'success'
            messages.success(request, "Result Detail has been deleted successfully.")
        except:
            resp['msg'] = 'result Detail has failed to delete.'

    return HttpResponse(json.dumps(resp),content_type="application/json")


def select_student(request):
    context['page'] = 'Select Student'
    context['page_title'] = 'Select Student'
    students = models.Student.objects.filter(status = 1).all()
    context['students'] = students

    return render(request, 'select_student_results.html', context)

def list_student_result(request, pk=None):
    if pk is None:
        messages.error(request, "Invalid Student ID")
        return redirect('login')
    else:
        student = models.Student.objects.get(id = pk)
        results = models.Result.objects.filter(student = student)
        context['student'] = student
        context['results'] = results
        context['page_title'] = str(student) + "'s Results"
        context['has_navigation'] = False
        context['has_sidebar'] = False

    return render(request, 'list_results.html', context)

# def add_subject_detail(request, pk=None):

#     report = SubjectDetail.objects.all()
#     if not report:
#         return render(request, 'record_not_found.html')

#     return render(request, 'add_subject_detail.html', {'reports': report})

def subjects(request):
    subjects = Subject.objects.all()
    return render(request, 'student_mgt.html', {'subjects': f"{subjects}"})
# --------------------------------------------------------------------------------------------#
def data_list(request):
    # under inspection
    # pdb.set_trace()
    class_students = Class.objects.all()
    return render(request, 'data_list.html', {'class_students': class_students})
# --------------------------------------------------------------------------------------------#
# front-debug-dropdown--> under inspection
# --------------------------------------------------------------------------------------------#
def student_class_list(request):
    # under inspection
    # pdb.set_trace()
    student_class = Class.objects.all()
    # debug->2
    # .values()
    # print(student_class)
    # return JsonResponse(list(student_class), safe=False)
    return render(request, 'student_mgt.html', {'student_class': student_class})
# --------------------------------------------------------------------------------------------#

def subjects_form_view(request):
    # under inspection
    classes = Class.objects.all()
    
    # Dictionary to store students for each class
    class_students = {}
    
    # Loop through classes and query students for each class
    for class_instance in classes:
        students = Student.objects.filter(student_class=class_instance)
        class_students[class_instance] = students
    
    return render(request, 'class_list.html', {'classes': classes, 'class_students': class_students})

@csrf_protect
def physics_form(request):

    # receives data from the post methods of the form

    if request.method == 'POST':
        # Extract data from the POST request
        #converting them to strings such that if the user leaves the  fields empty
        student_name = request.POST.get('student_name')
        subject_name = request.POST.get('subject_name')
        c1 = float(request.POST.get('c1')) if request.POST.get('c1') != '' else None
        c2 = float(request.POST.get('c2')) if request.POST.get('c2') != '' else None
        c3 = float(request.POST.get('c3')) if request.POST.get('c3') != '' else None
        c4 = float(request.POST.get('c4')) if request.POST.get('c4') != '' else None
        final = float(request.POST.get('final')) if request.POST.get('final') != '' else None
        
        # Validate if student exists
        try:
            student_instance = Student.objects.get(name=student_name)
        except Student.DoesNotExist:
            messages.error(request, "Student does not exist")
            # return HttpResponse("Student does not exist")

        # Validate if subject exists
        try:
            subject_instance = Subject.objects.get(name=subject_name)
        except Subject.DoesNotExist:
            messages.error(request, "Subject does not exist")
            # return HttpResponse("Subject does not exist")

        # Process form data and save it
        subject_detail, created = SubjectDetail.objects.get_or_create(
            student=student_instance,
            subject=subject_instance,
            defaults={'c1': c1, 'c2': c2, 'c3': c3, 'c4': c4, 'final': final}
        )

        messages.success(request,"Student records created successfully")
        if not created:
            # Update the existing record
            subject_detail.c1 = c1
            subject_detail.c2 = c2
            subject_detail.c3 = c3
            subject_detail.c4 = c4
            subject_detail.final = final
            subject_detail.save()

            messages.success(request,"Student records updated successfully")

        return redirect('student_mgt')
    else:
## ---------------------------------------debug-data-post-form---------------------------------------
##        print(request.POST.get('c1'))     ##
##        print(request.POST.get('c2'))     ##
##        print(request.POST.get('c3'))     ##
##        print(request.POST.get('c4'))     ##
##        print(request.POST.get('final'))  ##
# # ---------------------------------------debug-data-post-form---------------------------------------

        return HttpResponse("Invalid request method")



def copy_sheet_to_desktop(request):

    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    folder_name = "datasheets"
    folder_path = os.path.join(desktop_path, folder_name)
    
    # Create the folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Generate a random number for the filename
    random_number = random.randint(1000, 9999)
    file_name = f"datasheet_{random_number}.xlsx"
    file_path = os.path.join(folder_path, file_name)

    # Create a new workbook
    wb = Workbook()
    ws = wb.active  # Get the active worksheet

    # Define title for the sheet
    title = "Student Marks Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=44)  # Merge cells from A1 to AR1
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    data_align = Alignment(horizontal='center', vertical='center')

    # Define colors for subjects
    colors = ['FFC000', '00B0F0']

    # Define headers for student information
    ws.cell(row=2, column=1).value = "NAME"
    ws.cell(row=2, column=2).value = "LIN"
    ws.cell(row=2, column=3).value = "STREAM"
    ws.cell(row=2, column=4).value = "SEX"
    for col in range(1, 5):
        ws.cell(row=2, column=col).alignment = header_align
        ws.cell(row=2, column=col).font = header_font
        ws.cell(row=2, column=col).fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type='solid')

    for i in range(1, 11):
        subject_name = f"Subject {i}"
        start_col = 5 + (i - 1) * 4
        end_col = start_col + 3
        
        # Set alternating colors for subjects
        header_fill = PatternFill(start_color=colors[i % 2], end_color=colors[i % 2], fill_type='solid')
        
        # Merge cells for subject name and set styles
        ws.cell(row=2, column=start_col).value = subject_name
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(row=2, column=start_col).alignment = header_align
        ws.cell(row=2, column=start_col).fill = header_fill
        ws.cell(row=2, column=start_col).font = header_font

        # Set headers for each component of the subject
        ws.cell(row=3, column=start_col).value = "C1"
        ws.cell(row=3, column=start_col + 1).value = "C2"
        ws.cell(row=3, column=start_col + 2).value = "C3"
        ws.cell(row=3, column=start_col + 3).value = "FS"
        for col in range(start_col, end_col + 1):
            ws.cell(row=3, column=col).alignment = header_align
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font

    # Add borders to all cells in the sheet
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=44):
        for cell in row:
            cell.border = thin_border

    # Save the workbook as a spreadsheet file
    wb.save(file_path)
    messages.success(request, "Created successfully")

    return redirect('index')

    # messages.success(f"Spreadsheet created successfully and saved to desktop")


def filter_students(request):
    # if request.method == 'POST':
    selected_class_name = request.POST.get('class_filter')
    all_students = request.POST.get('all')

    filtered_students = Student.objects.all()

    if selected_class_name: 
        filtered_students = Student.objects.filter(classI__level=selected_class_name)
    elif all_students:
        filtered_students = Student.objects.all()
    else:
        filtered_students = Student.objects.all()

    subjects = Subject.objects.all()  
    students = Student.objects.all() 

    return render(request, 'student_mgt.html', {
        'filtered_students': filtered_students,
        'subjects': subjects,
        'students': students,
    })

def isolated_classes(request):
    isolated_classes = Class.objects.all()
    return render(request, 'clases.html', {'isolated_class': isolated_classes})

def return_home(request):
    # returning to the  previous page
    return redirect('student_mgt')

def view_sheet(request):
    # returning to the  previous page
    return redirect('view_sheet ')

def all_fields_filled_checker(request,subject_id, student_id):
    
    from django.db.models import Q

    subject= get_object_or_404(Subject, id=subject_id)
    student= get_object_or_404(Student, id=student_id)

    all_fields_filled = not SubjectDetail.objects.filter(
    student_id=student,
    subject_id=subject,
    ).filter(
        Q(c1__isnull=True) |
        Q(c2__isnull=True) |
        Q(c3__isnull=True) |
        Q(c4__isnull=True) |
        Q(final__isnull=True)
    ).exists()

    # return render(request, 'subject_mgt.html', {'all_fields_filled': all_fields_filled})

# ----------debug-resp-->db-queryset------------------------
    return HttpResponse(all_fields_filled)

# ---------------------------graph...................................#
#

    # Retrieve data dynamically (adjust based on your data source)
    subjects = ["Math", "Science", "English", "History"]  # Example with more subjects
    performance_ranges = ["High", "Average", "Low"]
    counts = [[8, 7, 5, 4], [6, 9, 5, 7], [7, 8, 5, 6]]  # Example with more counts
    colors = ["green", "yellow", "red"]
    bar_width = 0.25
    space_width = 0.1

    image_path = os.path.join(settings.MEDIA_ROOT, 'bar_chart.png')

    try:
        generate_chart_image(subjects, counts, colors, bar_width, space_width, image_path)

        with open(image_path, 'rb') as f:
            image_data = f.read()

        response = HttpResponse(image_data, content_type='image/png')
        return response

    except (ValueError, Exception) as e:
        # Handle errors gracefully, send appropriate response or error message
        return HttpResponse("Error generating chart: {}".format(e), status=500)

"""

start-uploads data from the sheet
student activities=> c1,c2,c3 and c4
filterling and insert data into the database

"""
def get_records():
    try:
        uploads_folder = os.path.join(settings.BASE_DIR, 'uploads')
        
        # Construct the absolute path to the file
        file_path = os.path.join(uploads_folder, 'card.xlsx')
        
        # Read data from the Excel file, skipping the first row for headers
        df = pd.read_excel(file_path, sheet_name='Sheet', skiprows=1)
        
        # Filter out columns with all NaN values
        df_filtered = df.dropna(axis=1, how='all')

        # Further skip the first two rows of data (originally the first three lines)
        df_filtered = df_filtered.iloc[1:]

        subjects = df.iloc[1]
        # Dropping the first four columns
        column_drop = subjects[3:]

        # Dropping subjects with unnamed head labels
        subject_modified = column_drop.filter(regex=r'^(?!Unnamed)')
        subject_modified = pd.DataFrame(subject_modified)

        # List to store records
        records = []

        print(f"Total rows to process: {len(df_filtered)}")  # Debugging output

        # Iterate through DataFrame rows 
        for index, row in df_filtered.iterrows():
            print(f"Processing row {index + 1}: {row}")  # Debugging output

            # Extract data from the row
            name = row['NAME']
            lin = row['LIN']
            stream = row['STREAM']
            sex = row['SEX']

            # Dictionary to store student data
            student_data = {'name': name, 'lin': lin, 'stream': stream, 'sex': sex, 'subjects': {}}

            # Iterate through subjects
            for subject in subject_modified.index:
                subject_data = {}

                # Get the index of the current subject column
                subject_index = df.columns.get_loc(subject)

                # Combine marks list and activities list
                activity_marks = []
                for i in range(subject_index, subject_index + 4):
                    cell_data = row.iloc[i]
                    if isinstance(cell_data, (int, float)):
                        activity_marks.append(round(cell_data, 1))
                    else:
                        activity_marks.append(cell_data)

                subject_data['activity_marks'] = activity_marks

                # Add subject data to the student's subjects dictionary
                student_data['subjects'][subject] = subject_data

            # Append student's data to the records list
            records.append(student_data)

        return records

    except FileNotFoundError:
        print("Error: File 'card.xlsx' not found.")
        return []
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return []

def excell_to_db(records=get_records()):
    if records:
        for idx, record in enumerate(records, start=1):
            print(f"Record {idx}:")
            print(f"Name: {record['name']}")
            for subject, data in record['subjects'].items():
                print(f"Subject: {subject}")
                print("Activities:")
                for idx, activity in enumerate(data['activity_marks'], start=1):
                    print(f"c{idx}: {activity}")
                print()
            print()
    else:
        print("No records found.")
        
# data test db variables
        
def excell_to_db_var(records=get_records()):
    data_list = []
    if records:
        for record in records:
            name = str(record['name'])
            lin=str(record['lin'])
            sex=str(record['sex'])
            stream=str(record['stream'])

            subjects_data = []
            for sub, data in record['subjects'].items():
                subject = sub
                activity_vars = {}
                for idx, activity in enumerate(data['activity_marks'], start=1):
                    activity_vars[f'c{idx}'] = activity
                subjects_data.append({'subject': subject, 'activity_vars': activity_vars})
            data_list.append({'name': name,'lin':lin, 'sex':sex, 'stream': stream, 'subjects_data': subjects_data})
    else:
        print("No records found.")
    
    return data_list


# call to db insert
def combine_records(records):
    combined_records = {}

    for record in records:
        name = record['name']
        lin = record['lin']
        gender = record['sex']
        stream = record['stream']

        if name not in combined_records:
            # Initialize a new combined record for this name
            combined_records[name] = {
                'name': name,
                'lin': lin,
                'gender': gender,
                'stream': stream,
                'subjects_data': {}
            }

        # Iterate through subjects in the current record
        for subject_data in record['subjects_data']:
            subject = subject_data['subject']
            activity_vars = subject_data['activity_vars']

            if subject not in combined_records[name]['subjects_data']:
                # Add subject and activities for this subject
                combined_records[name]['subjects_data'][subject] = activity_vars
            else:
                # Merge activities for existing subject
                combined_records[name]['subjects_data'][subject].update(activity_vars)

    return list(combined_records.values())


def student_record(request):
    try:
        records = excell_to_db_var()  # Returns raw records from Excel
        combined_records = combine_records(records)  # Combines records into desired format

        formatted_records = []
        
        for record in combined_records:
            name = record['name']
            lin = record['lin']
            gender = record['gender']
            stream = record['stream']

            subjects_data = {}
            for subject, activities in record['subjects_data'].items():
                subjects_data[subject] = {
                    'c1': activities['c1'],
                    'c2': activities['c2'],
                    'c3': activities['c3'],
                    'c4': activities['c4']
                }

            formatted_record = {
                'name': name,
                'lin': lin,
                'gender': gender,
                'stream': stream,
                'subjects_data': subjects_data
            }

            formatted_records.append(formatted_record)

        return formatted_records

    except Exception as e:
        print(f"An error occurred while formatting records: {str(e)}")
        return []


"""write to db"""

def write_to_db(request):
    try:
        combined_records = student_record(request)
    except Exception as e:
        print(f"Error getting student records: {str(e)}")
        return JsonResponse({'error': 'Failed to get student records', 'message': str(e)}, status=500)

    if not combined_records:
        print("No records found.")
        return JsonResponse({'message': 'No records found.'}, status=200)

    print(f"Total records to process: {len(combined_records)}")  # Debugging output

    for idx, record in enumerate(combined_records):
        try:
            print(f"Processing record {idx + 1}/{len(combined_records)}: {record}")  # Debugging output

            lin = int(float(record['lin']))  # Convert lin to float first, then to integer
            print("LIN:", lin)

            # Ensure 'defaults' does not overwrite existing students
            student, created = Student.objects.get_or_create(
                lin=lin,
                defaults={
                    'name': record['name'],
                    'gender': record['gender'],
                    'stream': record['stream']
                }
            )
            print(f"Student: {student}, Created: {created}")

            # Update existing student details if not created new
            if not created:
                student.name = record['name']
                student.gender = record['gender']
                student.stream = record['stream']
                student.save()

            for subject_name, activities in record['subjects_data'].items():
                print("Processing subject:", subject_name)  # Debugging output
                subject, created = Subject.objects.get_or_create(name=subject_name)
                print(f"Subject: {subject}, Created: {created}")

                student_subject, created = StudentSubject.objects.get_or_create(
                    student=student,
                    subject=subject,
                    defaults={
                        'c1': activities['c1'],
                        'c2': activities['c2'],
                        'c3': activities['c3'],
                        'fs': activities['c4']
                    }
                )
                print(f"StudentSubject: {student_subject}, Created: {created}")

                # Update existing student_subject details if not created new
                if not created:
                    student_subject.c1 = activities['c1']
                    student_subject.c2 = activities['c2']
                    student_subject.c3 = activities['c3']
                    student_subject.fs = activities['c4']
                    student_subject.save()

                for idx, activity in enumerate([activities['c1'], activities['c2'], activities['c3'], activities['c4']], start=1):
                    print(f"Processing activity c{idx}: {activity}")  # Debugging output
                    mark, created = Marks.objects.get_or_create(
                        student_subject=student_subject,
                        component_name=f'c{idx}',
                        defaults={'marks': activity}
                    )
                    print(f"Mark: {mark}, Created: {created}")

                    # Update existing marks details if not created new
                    if not created:
                        mark.marks = activity
                        mark.save()

        except Exception as e:
            print(f"Error processing record {idx + 1}/{len(combined_records)}: {record}, Error: {str(e)}")  # Debugging output
            continue  # Skip this record and move to the next

    print("Data successfully written to the database.")
    print(student_record(request))
    print("test records: ", get_records())
    return JsonResponse({'message': 'Data successfully written to the database.'}, status=200)

@csrf_protect
def upload_file(request):
    """
    uploading the file for excell file
    """
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        if uploaded_file:
            upload_folder = 'uploads'
            
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            for filename in os.listdir(upload_folder):
                    file_path = os.path.join(upload_folder, filename)
                    try:
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                    except Exception as e:
                        print(e)
            upload_path = os.path.join(upload_folder, uploaded_file.name)
            
            with open(upload_path, 'wb') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            messages.success(request, "File uploaded successfully.")
            return redirect('upload_success')
        else:
            messages.error(request, "No file was uploaded.")
            
    
    return render(request, 'includes/navigation.html')



def upload_success(request):
    return render(request, 'pages/index.html')

def create_or_update_school_details(request):
    if request.method == 'POST':
        form = SchoolDetailsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('school_details_success')
    else:
        form = SchoolDetailsForm()
    
    return render(request, 'school_details_form.html', {'form': form})

def school_details_success(request):
    return render(request, 'school_details_success.html')

def render_report_data(request):
    pass